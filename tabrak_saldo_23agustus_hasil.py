"""
Script: tabrak_saldo_23agustus_hasil.py
Tujuan: Mencocokkan CIFNO dari file
        HASIL_TABRAK_SALDO_KECIL_FEB26_APRIL30_7MEI_17MEI.xlsx (kolom 19)
        dengan CIFNO di file-file CSV DI319_23AGUS2026,
        lalu mengisi kolom 'saldo 23 agustus' (kolom 46).
"""

import pandas as pd
import openpyxl
import os
import glob
import time


def main():
    start_time = time.time()

    excel_path = r"C:\bestline\HASIL_TABRAK_SALDO_KECIL_FEB26_APRIL30_7MEI_17MEI.xlsx"
    csv_dir = r"C:\DI319\DI319_23AGUS2026"
    CIFNO_COL = 19      # Kolom S = CIFNO (uppercase)
    SALDO_COL = 46      # Kolom AT = saldo 23 agustus

    # ─── TAHAP 1: Membaca File Excel ─────────────────────────────────────────
    print("=" * 60)
    print("  TABRAK SALDO 23 AGUSTUS – HASIL_TABRAK VIA CIFNO")
    print("=" * 60)

    print("\n=== TAHAP 1: Membaca File Excel ===")
    if not os.path.exists(excel_path):
        print(f"Error: File Excel tidak ditemukan di {excel_path}")
        return

    wb = openpyxl.load_workbook(excel_path)
    sheet = wb.active
    print(f"Active Sheet: '{sheet.title}'")
    print(f"Max rows: {sheet.max_row}, Max cols: {sheet.max_column}")

    # Header check / set
    header_46 = sheet.cell(row=1, column=SALDO_COL).value
    print(f"Header kolom {SALDO_COL} saat ini: '{header_46}'")
    if header_46 != "saldo 23 agustus":
        sheet.cell(row=1, column=SALDO_COL).value = "saldo 23 agustus"
        print(f"Header kolom {SALDO_COL} diset ke: 'saldo 23 agustus'")

    # Kumpulkan target CIFNO dari kolom 19, mulai dari baris 2
    target_cifnos = set()
    row_mapping = {}  # cifno_str -> list of row indices

    total_excel_rows = sheet.max_row
    non_null_count = 0

    for row in range(2, total_excel_rows + 1):
        cell_val = sheet.cell(row=row, column=CIFNO_COL).value
        if cell_val is not None:
            cifno_str = str(cell_val).strip().upper()
            if cifno_str and cifno_str != "NAN" and cifno_str != "NONE":
                target_cifnos.add(cifno_str)
                if cifno_str not in row_mapping:
                    row_mapping[cifno_str] = []
                row_mapping[cifno_str].append(row)
                non_null_count += 1

    print(f"Total baris di Excel: {total_excel_rows - 1}")
    print(f"Total CIFNO terisi: {non_null_count}")
    print(f"Total CIFNO unik: {len(target_cifnos)}")

    # ─── TAHAP 2: Baca CSV DI319 dan bangun dict CIFNO -> total Saldo ──
    print("\n=== TAHAP 2: Membaca CSV DI319 23 Agustus ===")
    csv_files = glob.glob(os.path.join(csv_dir, "*.csv"))
    print(f"Ditemukan {len(csv_files)} file CSV di {csv_dir}.")

    cif_balance = {}   # CIFNO (str) -> total BALANCE (float)
    total_csv_rows = 0

    for idx, fpath in enumerate(csv_files, 1):
        filename = os.path.basename(fpath)
        file_start = time.time()
        print(f"[{idx}/{len(csv_files)}] Membaca {filename} ...", end=" ", flush=True)

        file_matches = 0
        try:
            chunk_count = 0
            for chunk in pd.read_csv(
                fpath,
                skiprows=3,
                chunksize=200000,
                usecols=["CIFNO", "BALANCE"],
                dtype={"CIFNO": str, "BALANCE": str},
                on_bad_lines="skip",
                encoding="latin1",
            ):
                chunk_count += 1

                # Clean CIFNO
                chunk["CIFNO"] = chunk["CIFNO"].str.strip().str.upper()

                # Clean BALANCE: remove commas, convert to numeric
                chunk["BALANCE"] = (
                    chunk["BALANCE"]
                    .str.replace(",", "", regex=False)
                    .str.strip()
                )
                chunk["BALANCE"] = pd.to_numeric(chunk["BALANCE"], errors="coerce").fillna(0)

                total_csv_rows += len(chunk)

                # Filter hanya CIFNO yang ada di target
                matches = chunk[chunk["CIFNO"].isin(target_cifnos)]
                if not matches.empty:
                    grp = matches.groupby("CIFNO")["BALANCE"].sum()
                    for cifno, saldo in grp.items():
                        cif_balance[cifno] = cif_balance.get(cifno, 0) + saldo
                        file_matches += 1

            elapsed = time.time() - file_start
            print(f"OK ({file_matches} CIF match, {chunk_count} chunks, {elapsed:.1f}s)")

        except Exception as e:
            print(f"ERROR: {e}")

    print(f"\nTotal baris CSV dibaca: {total_csv_rows:,}")
    print(f"Total CIF unik yang match: {len(cif_balance):,}")

    # ─── TAHAP 3: Mengisi Kolom Saldo 23 Agustus di Excel ────────────────────────
    print(f"\n=== TAHAP 3: Mengisi Kolom 'saldo 23 agustus' (Kolom {SALDO_COL}) ===")
    matched_rows_count = 0
    total_saldo_filled = 0.0

    for cifno, saldo in cif_balance.items():
        rows = row_mapping.get(cifno, [])
        for r in rows:
            sheet.cell(row=r, column=SALDO_COL).value = saldo
            matched_rows_count += 1
            total_saldo_filled += saldo

    print(f"Berhasil mengupdate {matched_rows_count} baris di kolom {SALDO_COL}.")
    print(f"Total saldo terisi: Rp {total_saldo_filled:,.2f}")

    # Hitung yang tidak match
    unmatched = target_cifnos - set(cif_balance.keys())
    print(f"Total CIFNO tidak match: {len(unmatched)}")

    # ─── TAHAP 4: Menyimpan File Excel ───────────────────────────────────────
    print("\n=== TAHAP 4: Menyimpan File Excel ===")
    print(f"Menyimpan perubahan ke {os.path.basename(excel_path)} ...", end=" ", flush=True)
    save_start = time.time()
    wb.save(excel_path)
    print(f"SUKSES! ({time.time() - save_start:.2f}s)")

    duration = time.time() - start_time
    print(f"\n{'=' * 60}")
    print(f"  Proses selesai dalam {duration / 60:.2f} menit ({duration:.2f} detik)")
    print(f"{'=' * 60}")


if __name__ == "__main__":
    main()
