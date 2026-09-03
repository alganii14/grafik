"""
Script: tabrak_saldo_23agustus_ppatk.py
Tujuan: Mencocokkan norek (acctno) dari file
        MASTER PPATK DAN PENYANGGA_APRIL30_7MEI_17MEI.xlsx (kolom 5)
        dengan No_Rekening (textbox15) di file-file CSV DI319_23AGUS2026,
        lalu mengisi kolom 'saldo 23 agustus' (kolom 40).
"""

import pandas as pd
import openpyxl
import os
import glob
import time


def main():
    start_time = time.time()

    excel_path = r"C:\bestline\MASTER PPATK DAN PENYANGGA_APRIL30_7MEI_17MEI.xlsx"
    csv_dir = r"C:\DI319\DI319_23AGUS2026"
    NOREK_COL = 5       # Kolom E = acctno
    SALDO_COL = 40      # Kolom AN = saldo 23 agustus

    # ─── TAHAP 1: Membaca File Excel ─────────────────────────────────────────
    print("=" * 60)
    print("  TABRAK SALDO 23 AGUSTUS – PPATK VIA NOREK")
    print("=" * 60)

    print("\n=== TAHAP 1: Membaca File Excel ===")
    if not os.path.exists(excel_path):
        print(f"Error: File Excel tidak ditemukan di {excel_path}")
        return

    wb = openpyxl.load_workbook(excel_path)
    sheet = wb.active
    print(f"Active Sheet: '{sheet.title}'")
    print(f"Max rows: {sheet.max_row}, Max cols: {sheet.max_column}")

    # Header check
    header_40 = sheet.cell(row=1, column=SALDO_COL).value
    print(f"Header kolom {SALDO_COL} saat ini: '{header_40}'")
    if header_40 != "saldo 23 agustus":
        sheet.cell(row=1, column=SALDO_COL).value = "saldo 23 agustus"
        print(f"Header kolom {SALDO_COL} diset ke: 'saldo 23 agustus'")

    # Kumpulkan target norek dari kolom 5, mulai dari baris 2
    target_noreks = set()
    row_mapping = {}  # norek_str -> list of row indices

    total_excel_rows = sheet.max_row
    non_null_count = 0

    for row in range(2, total_excel_rows + 1):
        cell_val = sheet.cell(row=row, column=NOREK_COL).value
        if cell_val is not None:
            norek_str = str(cell_val).strip()
            # Jika berupa float (misal 3.37011e+13), konversi ke int dulu
            try:
                norek_str = str(int(float(norek_str)))
            except (ValueError, OverflowError):
                pass
            if norek_str and norek_str != "NAN" and norek_str != "NONE":
                target_noreks.add(norek_str)
                if norek_str not in row_mapping:
                    row_mapping[norek_str] = []
                row_mapping[norek_str].append(row)
                non_null_count += 1

    print(f"Total baris di Excel: {total_excel_rows - 1}")
    print(f"Total norek terisi: {non_null_count}")
    print(f"Total norek unik: {len(target_noreks)}")

    # ─── TAHAP 2: Baca CSV DI319 dan bangun dict norek -> Saldo ─────────
    print("\n=== TAHAP 2: Membaca CSV DI319 23 Agustus ===")
    csv_files = glob.glob(os.path.join(csv_dir, "*.csv"))
    print(f"Ditemukan {len(csv_files)} file CSV di {csv_dir}.")

    norek_balance = {}   # norek (str) -> total BALANCE (float)
    total_csv_rows = 0

    for idx, fpath in enumerate(csv_files, 1):
        filename = os.path.basename(fpath)
        file_start = time.time()
        print(f"[{idx}/{len(csv_files)}] Membaca {filename} ...", end=" ", flush=True)

        file_matches = 0
        try:
            chunk_count = 0
            for chunk in pd.read_csv(
                fpath,
                skiprows=3,
                chunksize=200000,
                usecols=["textbox15", "BALANCE"],
                dtype={"textbox15": str, "BALANCE": str},
                on_bad_lines="skip",
                encoding="latin1",
            ):
                chunk_count += 1

                # Clean textbox15 (No_Rekening)
                chunk["textbox15"] = chunk["textbox15"].str.strip()

                # Clean BALANCE
                chunk["BALANCE"] = (
                    chunk["BALANCE"]
                    .str.replace(",", "", regex=False)
                    .str.strip()
                )
                chunk["BALANCE"] = pd.to_numeric(chunk["BALANCE"], errors="coerce").fillna(0)

                total_csv_rows += len(chunk)

                # Filter hanya norek yang ada di target
                matches = chunk[chunk["textbox15"].isin(target_noreks)]
                if not matches.empty:
                    grp = matches.groupby("textbox15")["BALANCE"].sum()
                    for norek, saldo in grp.items():
                        norek_balance[norek] = norek_balance.get(norek, 0) + saldo
                        file_matches += 1

            elapsed = time.time() - file_start
            print(f"OK ({file_matches} rek match, {chunk_count} chunks, {elapsed:.1f}s)")

        except Exception as e:
            print(f"ERROR: {e}")

    print(f"\nTotal baris CSV dibaca: {total_csv_rows:,}")
    print(f"Total norek unik yang match: {len(norek_balance):,}")

    # ─── TAHAP 3: Mengisi Kolom Saldo 23 Agustus di Excel ────────────────────────
    print(f"\n=== TAHAP 3: Mengisi Kolom 'saldo 23 agustus' (Kolom {SALDO_COL}) ===")
    matched_rows_count = 0
    total_saldo_filled = 0.0

    for norek, saldo in norek_balance.items():
        rows = row_mapping.get(norek, [])
        for r in rows:
            sheet.cell(row=r, column=SALDO_COL).value = saldo
            matched_rows_count += 1
            total_saldo_filled += saldo

    print(f"Berhasil mengupdate {matched_rows_count} baris di kolom {SALDO_COL}.")
    print(f"Total saldo terisi: Rp {total_saldo_filled:,.2f}")

    # Hitung yang tidak match
    unmatched = target_noreks - set(norek_balance.keys())
    print(f"Total norek tidak match: {len(unmatched)}")

    # ─── TAHAP 4: Menyimpan File Excel ───────────────────────────────────────
    print("\n=== TAHAP 4: Menyimpan File Excel ===")
    print(f"Menyimpan perubahan ke {os.path.basename(excel_path)} ...", end=" ", flush=True)
    save_start = time.time()
    wb.save(excel_path)
    print(f"SUKSES! ({time.time() - save_start:.2f}s)")

    duration = time.time() - start_time
    print(f"\n{'=' * 60}")
    print(f"  Proses selesai dalam {duration / 60:.2f} menit ({duration:.2f} detik)")
    print(f"{'=' * 60}")


if __name__ == "__main__":
    main()
